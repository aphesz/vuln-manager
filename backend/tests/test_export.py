"""
Tests for export endpoint with CSV, Excel, column selection, and filtering.
"""
import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session, select
from app.main import app
from app.models import Project, Finding, Instance
from app.timezone_utils import get_utc_now
import io
from openpyxl import load_workbook
import csv


@pytest.fixture
def sample_project(session: Session):
    """Create a sample project with varied findings for export testing."""
    project = Project(name="Export Test Project", consultant_name="Tester")
    session.add(project)
    session.commit()
    session.refresh(project)
    
    # Create findings with different risk levels and statuses
    findings_data = [
        {
            "title": "Critical SQL Injection",
            "risk_rating": "Critical",
            "description": "SQL injection in login form",
            "remediation": "Use parameterized queries",
            "review_status": "Approved",
            "reviewer_name": "Alice",
            "issue_status": "Open",
            "jira_issue_key": "VULN-101",
        },
        {
            "title": "High XSS Vulnerability",
            "risk_rating": "High",
            "description": "Reflected XSS in search",
            "remediation": "Sanitize user input",
            "review_status": "In Review",
            "reviewer_name": "Bob",
            "issue_status": "Open",
            "jira_issue_key": "VULN-102",
        },
        {
            "title": "Medium Info Disclosure",
            "risk_rating": "Medium",
            "description": "Sensitive data in error messages",
            "remediation": "Generic error messages",
            "review_status": "Pending",
            "issue_status": "Partially Closed",
        },
        {
            "title": "Low Missing Headers",
            "risk_rating": "Low",
            "description": "Missing security headers",
            "remediation": "Add security headers",
            "review_status": "Rejected",
            "issue_status": "Closed",
        },
    ]
    
    for data in findings_data:
        finding = Finding(project_id=project.id, **data)
        session.add(finding)
        session.commit()
        session.refresh(finding)
        
        # Add some instances
        for i in range(2):
            instance = Instance(
                finding_id=finding.id,
                location=f"/endpoint/{i}",
                details=f"Instance {i}",
                status="Open",
                created_at=get_utc_now()
            )
            session.add(instance)
    
    session.commit()
    return project


def test_export_excel_default(client: TestClient, sample_project: Project):
    """Test default Excel export with all columns and no filters."""
    response = client.get(f"/projects/{sample_project.id}/export?format=excel")
    
    assert response.status_code == 200
    assert response.headers['content-type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'attachment' in response.headers.get('content-disposition', '')
    
    # Parse Excel file
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check headers (all columns should be present)
    headers = [cell.value for cell in ws[1]]
    assert 'Title' in headers
    assert 'Risk Rating' in headers
    assert 'Description' in headers
    assert 'Remediation' in headers
    assert 'Instance Count' in headers
    
    # Check data rows (4 findings)
    assert ws.max_row == 5  # Header + 4 findings
    
    # Verify instance counts
    instance_col_idx = headers.index('Instance Count') + 1
    for row in range(2, 6):  # Rows 2-5
        assert ws.cell(row, instance_col_idx).value == 2  # Each finding has 2 instances


def test_export_csv_default(client: TestClient, sample_project: Project):
    """Test default CSV export."""
    response = client.get(f"/projects/{sample_project.id}/export?format=csv")
    
    assert response.status_code == 200
    assert 'text/csv' in response.headers['content-type']
    assert 'attachment' in response.headers.get('content-disposition', '')
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    assert len(rows) == 4  # 4 findings
    assert 'Title' in reader.fieldnames
    assert 'Risk Rating' in reader.fieldnames


def test_export_with_column_selection(client: TestClient, sample_project: Project):
    """Test export with specific columns selected."""
    columns = "title,risk_rating,instance_count"
    response = client.get(
        f"/projects/{sample_project.id}/export?format=excel&columns={columns}"
    )
    
    assert response.status_code == 200
    
    # Parse Excel file
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check headers (only selected columns)
    headers = [cell.value for cell in ws[1]]
    assert headers == ['Title', 'Risk Rating', 'Instance Count']
    assert 'Description' not in headers
    assert 'Remediation' not in headers


def test_export_with_risk_filter(client: TestClient, sample_project: Project):
    """Test export with risk level filtering."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=csv&risk_filter=Critical,High"
    )
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # Should only have 2 findings (Critical and High)
    assert len(rows) == 2
    risk_levels = {row['Risk Rating'] for row in rows}
    assert risk_levels == {'Critical', 'High'}


def test_export_with_issue_status_filter(client: TestClient, sample_project: Project):
    """Test export with issue status filtering."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=csv&status_filter=Open"
    )
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # Should only have 2 findings with Open status
    assert len(rows) == 2


def test_export_with_review_status_filter(client: TestClient, sample_project: Project):
    """Test export with review status filtering."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=csv&review_filter=Approved,In Review"
    )
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # Should only have 2 findings (Approved and In Review)
    assert len(rows) == 2


def test_export_with_multiple_filters(client: TestClient, sample_project: Project):
    """Test export with multiple filters combined."""
    response = client.get(
        f"/projects/{sample_project.id}/export?"
        f"format=csv&risk_filter=Critical,High&review_filter=Approved"
    )
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # Should only have 1 finding (Critical + Approved)
    assert len(rows) == 1
    assert rows[0]['Risk Rating'] == 'Critical'
    assert rows[0]['Review Status'] == 'Approved'


def test_export_with_column_selection_and_filters(client: TestClient, sample_project: Project):
    """Test export with both column selection and filters."""
    response = client.get(
        f"/projects/{sample_project.id}/export?"
        f"format=excel&columns=title,risk_rating&risk_filter=High,Medium"
    )
    
    assert response.status_code == 200
    
    # Parse Excel file
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check headers (only selected columns)
    headers = [cell.value for cell in ws[1]]
    assert headers == ['Title', 'Risk Rating']
    
    # Check row count (2 findings: High and Medium)
    assert ws.max_row == 3  # Header + 2 findings


def test_export_invalid_format(client: TestClient, sample_project: Project):
    """Test export with invalid format."""
    response = client.get(f"/projects/{sample_project.id}/export?format=pdf")
    
    assert response.status_code == 400
    assert "Invalid format" in response.json()['detail']


def test_export_invalid_columns(client: TestClient, sample_project: Project):
    """Test export with invalid column names."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=excel&columns=title,invalid_column"
    )
    
    assert response.status_code == 400
    assert "Invalid columns" in response.json()['detail']


def test_export_nonexistent_project(client: TestClient):
    """Test export for non-existent project."""
    response = client.get("/projects/99999/export?format=excel")
    
    assert response.status_code == 404
    assert "Project not found" in response.json()['detail']


def test_export_empty_project(client: TestClient, session: Session):
    """Test export for project with no findings."""
    project = Project(name="Empty Project", consultant_name="Tester")
    session.add(project)
    session.commit()
    session.refresh(project)
    
    response = client.get(f"/projects/{project.id}/export?format=csv")
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # Should have headers but no data rows
    assert len(rows) == 0
    assert len(reader.fieldnames) > 0  # Headers present


def test_export_filters_no_matches(client: TestClient, sample_project: Project):
    """Test export with filters that match no findings."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=csv&risk_filter=Informational"
    )
    
    assert response.status_code == 200
    
    # Parse CSV
    content = response.content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    
    # No findings should match
    assert len(rows) == 0


def test_export_json_format(client: TestClient, sample_project: Project):
    """Test JSON export with full data structure."""
    response = client.get(f"/projects/{sample_project.id}/export?format=json")
    
    assert response.status_code == 200
    assert response.headers['content-type'] == 'application/json'
    
    data = response.json()
    
    # Check structure
    assert 'project' in data
    assert 'export_metadata' in data
    assert 'findings' in data
    
    # Check project metadata
    assert data['project']['id'] == sample_project.id
    assert data['project']['name'] == sample_project.name
    
    # Check export metadata
    assert 'exported_at' in data['export_metadata']
    assert data['export_metadata']['total_findings'] > 0
    assert 'columns_included' in data['export_metadata']
    
    # Check findings data
    assert len(data['findings']) > 0
    finding = data['findings'][0]
    assert 'title' in finding
    assert 'risk_rating' in finding


def test_export_json_with_filters(client: TestClient, sample_project: Project):
    """Test JSON export with risk filter."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=json&risk_filter=Critical"
    )
    
    assert response.status_code == 200
    data = response.json()
    
    # Check filters applied
    assert data['export_metadata']['filters_applied']['risk_levels'] == ['Critical']
    
    # All findings should be Critical
    for finding in data['findings']:
        assert finding['risk_rating'] == 'Critical'


def test_export_json_with_column_selection(client: TestClient, sample_project: Project):
    """Test JSON export with selected columns only."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=json&columns=title,risk_rating"
    )
    
    assert response.status_code == 200
    data = response.json()
    
    # Check only selected columns present
    if data['findings']:
        finding = data['findings'][0]
        assert 'title' in finding
        assert 'risk_rating' in finding
        # Other columns should not be present
        assert 'description' not in finding or finding['description'] is None


def test_export_markdown_format(client: TestClient, sample_project: Project):
    """Test Markdown export format."""
    response = client.get(f"/projects/{sample_project.id}/export?format=markdown")
    
    assert response.status_code == 200
    assert 'text/markdown' in response.headers['content-type']
    
    content = response.content.decode('utf-8')
    
    # Check markdown structure
    assert f"# Vulnerability Assessment Report: {sample_project.name}" in content
    assert f"**Consultant:** {sample_project.consultant_name}" in content
    assert "## Summary" in content
    assert "## Findings" in content
    
    # Check for risk level table
    assert "| Risk Level | Count |" in content
    
    # Check for findings sections
    assert "### 1." in content  # First finding
    assert "**Risk Rating:**" in content


def test_export_markdown_with_instances(client: TestClient, sample_project: Project, session: Session):
    """Test Markdown export includes instance details."""
    # Add instances to a finding
    finding = session.exec(
        select(Finding).where(Finding.project_id == sample_project.id)
    ).first()
    
    instance1 = Instance(
        finding_id=finding.id,
        url="https://example.com/page1",
        parameter="username",
        evidence="<script>alert(1)</script>"
    )
    instance2 = Instance(
        finding_id=finding.id,
        url="https://example.com/page2",
        parameter="search",
        evidence="' OR '1'='1"
    )
    session.add(instance1)
    session.add(instance2)
    session.commit()
    
    response = client.get(
        f"/projects/{sample_project.id}/export?format=markdown&columns=title,risk_rating,instance_count"
    )
    
    assert response.status_code == 200
    content = response.content.decode('utf-8')
    
    # Check instance details
    assert "**Instances Found:**" in content
    assert "**Instance Details:**" in content
    assert "https://example.com/page1" in content
    assert "https://example.com/page2" in content
    assert "username" in content
    assert "search" in content


def test_export_markdown_with_filters(client: TestClient, sample_project: Project):
    """Test Markdown export with risk filter."""
    response = client.get(
        f"/projects/{sample_project.id}/export?format=markdown&risk_filter=Critical,High"
    )
    
    assert response.status_code == 200
    content = response.content.decode('utf-8')
    
    # Should only contain Critical and High findings
    assert "🔴 Critical" in content or "🟠 High" in content
    # Should not contain Medium or Low
    assert "🟡 Medium" not in content or content.count("### ") <= 2


def test_export_markdown_emoji_risk_badges(client: TestClient, sample_project: Project):
    """Test Markdown export includes emoji risk badges."""
    response = client.get(f"/projects/{sample_project.id}/export?format=markdown")
    
    assert response.status_code == 200
    content = response.content.decode('utf-8')
    
    # Check for risk emoji badges
    risk_emojis = {
        'Critical': '🔴',
        'High': '🟠',
        'Medium': '🟡',
        'Low': '🟢',
        'Informational': '🔵'
    }
    
    # At least one emoji should be present
    found_emoji = False
    for emoji in risk_emojis.values():
        if emoji in content:
            found_emoji = True
            break
    assert found_emoji


def test_export_format_filenames(client: TestClient, sample_project: Project):
    """Test that different formats have correct file extensions."""
    formats = {
        'excel': '.xlsx',
        'csv': '.csv',
        'json': '.json',
        'markdown': '.md'
    }
    
    for format_type, extension in formats.items():
        response = client.get(f"/projects/{sample_project.id}/export?format={format_type}")
        assert response.status_code == 200
        
        content_disposition = response.headers.get('content-disposition', '')
        assert extension in content_disposition
